"""Un fichero de Excel por empleado, con una hoja por mes, en OneDrive.

La organizacion importa: como el sueldo por hora se escribe a mano, tenerlo en
un sitio por persona y no por mes evita volver a teclear 49 precios cada primero
de mes. Vive en la hoja "Ficha", y las hojas de cada mes lo referencian.

Por eso cada ejecucion no genera el libro de cero: descarga el que ya existe,
le anade o reemplaza la hoja del mes y lo vuelve a subir. Asi se conserva lo que
haya escrito una persona: el precio hora y cualquier correccion a mano.
"""

import io
import logging
import re
import time

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

GRAPH = "https://graph.microsoft.com/v1.0"
TIMEOUT = 120

CABECERAS = [("fecha", 12), ("entrada", 10), ("salida", 10), ("horas", 9), ("aviso", 46)]

AMARILLO = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
# Azul claro para lo que rellena una persona: el amarillo ya significa
# "este dia hay que revisarlo" y no conviene mezclar los dos sentidos.
AZUL_EDITABLE = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
BORDE = Border(*[Side(style="thin", color="9BC2E6")] * 4)

EUROS = '#,##0.00 "€"'

HOJA_FICHA = "Ficha"
# Celdas que se rellenan a mano en la Ficha.
FILA_SUELDO = 6
FILA_TIPO = 7
FILA_HORAS_CONTRATO = 8

FILA_TABLA_MESES = 11      # cabecera de la tabla de meses en la Ficha
FILA_LIMPIEZA = 9          # desde donde se borra al rehacer la tabla; las
                           # primeras versiones la tenian en la 9
PRIMERA_FILA_DATOS = 10    # en la hoja de un mes: 1 titulo, 3-8 totales, 9 cabeceras

PAGO_POR_HORAS = "horas"
PAGO_POR_CONTRATO = "contrato"

PROHIBIDOS_FICHERO = re.compile(r'[<>:"/\\|?*]')

MESES = ["enero", "febrero", "marzo", "abril", "mayo", "junio",
         "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre"]
ORDEN_MES = {m: i for i, m in enumerate(MESES)}
PERIODO_ANTIGUO = re.compile(r"^(\d{4})-(\d{2})$")


def nombre_de_mes(periodo):
    """'2026-08' -> 'agosto'."""
    return MESES[int(periodo[5:7]) - 1]


def _orden_de_hoja(nombre):
    """Ordena las hojas por mes natural, no alfabeticamente.

    Sin esto 'agosto' iria antes que 'enero', que es justo lo que no queremos.
    """
    return ORDEN_MES.get(nombre, 99)


def _migrar_nombres_antiguos(libro):
    """Renombra las hojas 'YYYY-MM' de la primera version a nombre de mes.

    Sin esto, al anadir 'agosto' quedaria junto a la vieja '2026-08' y el mes
    apareceria dos veces en la ficha.
    """
    for hoja in list(libro.worksheets):
        m = PERIODO_ANTIGUO.match(hoja.title)
        if not m:
            continue
        nuevo = MESES[int(m.group(2)) - 1]
        if nuevo in libro.sheetnames:
            del libro[hoja.title]
        else:
            hoja.title = nuevo


def nombre_de_fichero(empleado, workno):
    """'Ana Canales - 54.xlsx'. Ordena por nombre y el numero lo hace unico."""
    base = PROHIBIDOS_FICHERO.sub(" ", empleado or "").strip() or "Sin nombre"
    base = " ".join(base.split())
    return "{} - {}.xlsx".format(base[:80], workno)


def token(tenant_id, client_id, client_secret):
    respuesta = requests.post(
        "https://login.microsoftonline.com/{}/oauth2/v2.0/token".format(tenant_id),
        data={
            "grant_type": "client_credentials",
            "client_id": client_id,
            "client_secret": client_secret,
            "scope": "https://graph.microsoft.com/.default",
        },
        timeout=TIMEOUT,
    )
    respuesta.raise_for_status()
    return respuesta.json()["access_token"]


# --- Graph -----------------------------------------------------------------

class Bloqueado(RuntimeError):
    """El fichero esta abierto en Excel y OneDrive no deja escribirlo."""


def _cabeceras(tk, tipo=None):
    h = {"Authorization": "Bearer {}".format(tk)}
    if tipo:
        h["Content-Type"] = tipo
    return h


def descargar(tk, drive_id, carpeta_id, nombre):
    """Contenido del fichero, o None si todavia no existe."""
    url = "{}/drives/{}/items/{}:/{}:/content".format(GRAPH, drive_id, carpeta_id, nombre)
    respuesta = requests.get(url, headers=_cabeceras(tk), timeout=TIMEOUT)
    if respuesta.status_code == 404:
        return None
    respuesta.raise_for_status()
    return respuesta.content


def subir(tk, drive_id, carpeta_id, nombre, contenido):
    """Sube o reemplaza el fichero.

    Si alguien lo tiene abierto en Excel, OneDrive responde 423 y aqui se sale
    enseguida con Bloqueado, sin esperar. Antes se reintentaba dos veces con
    veinte segundos de pausa y eso, con muchos ficheros abiertos a la vez,
    llegaria a agotar los diez minutos de la funcion. Quien llama junta los
    bloqueados y les da una segunda pasada al final.
    """
    url = "{}/drives/{}/items/{}:/{}:/content".format(GRAPH, drive_id, carpeta_id, nombre)
    tipo = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    respuesta = requests.put(url, headers=_cabeceras(tk, tipo), data=contenido, timeout=TIMEOUT)

    if respuesta.status_code in (200, 201):
        datos = respuesta.json()
        logging.info("Subido %s (%s bytes)", datos.get("name"), datos.get("size"))
        return datos.get("webUrl")

    if respuesta.status_code == 423:
        raise Bloqueado("{} esta abierto en Excel".format(nombre))

    raise RuntimeError("Graph rechazo la subida de {}: {} {}".format(
        nombre, respuesta.status_code, respuesta.text[:250]))


# --- Construccion del libro ------------------------------------------------

def _ficha(libro, empleado, workno, departamento):
    """Crea o completa la hoja Ficha. Nunca pisa lo que haya escrito una persona.

    Las etiquetas y el desplegable se reponen en cada ejecucion, para que los
    ficheros creados por versiones anteriores acaben teniendo tambien el
    selector de tipo de pago. Los valores solo se rellenan si estan vacios.
    """
    nueva = HOJA_FICHA not in libro.sheetnames
    hoja = libro.create_sheet(HOJA_FICHA, 0) if nueva else libro[HOJA_FICHA]

    for fila, etiqueta, pista in (
        (FILA_SUELDO, "SUELDO POR HORA",
         "← se escribe una sola vez, vale para todos los meses"),
        (FILA_TIPO, "TIPO DE PAGO",
         "← 'horas' paga todas las horas; 'contrato' paga solo las extra"),
        (FILA_HORAS_CONTRATO, "HORAS POR CONTRATO",
         "← horas mensuales del contrato (solo si es por contrato)"),
    ):
        hoja.cell(row=fila, column=1, value=etiqueta).font = Font(bold=True)
        celda = hoja.cell(row=fila, column=2)
        celda.fill = AZUL_EDITABLE
        celda.border = BORDE
        hoja.cell(row=fila, column=3, value=pista).font = Font(italic=True, color="808080")

    hoja.cell(row=FILA_SUELDO, column=2).number_format = EUROS
    hoja.cell(row=FILA_HORAS_CONTRATO, column=2).number_format = "0.00"

    validacion = DataValidation(
        type="list",
        formula1='"{},{}"'.format(PAGO_POR_HORAS, PAGO_POR_CONTRATO),
        allow_blank=True,
    )
    hoja.add_data_validation(validacion)
    celda_tipo = hoja.cell(row=FILA_TIPO, column=2)
    validacion.add(celda_tipo)
    # Por defecto 'horas', que es como venia funcionando: asi solo hay que tocar
    # la ficha de quien cobre por contrato.
    if celda_tipo.value is None:
        celda_tipo.value = PAGO_POR_HORAS

    hoja["A1"] = empleado
    hoja["A1"].font = Font(bold=True, size=14)
    hoja["A3"] = "Nº de empleado"
    hoja["B3"] = workno
    hoja["A4"] = "Departamento"
    hoja["B4"] = departamento
    hoja.column_dimensions["A"].width = 22
    hoja.column_dimensions["B"].width = 16
    for col in ("C", "D", "E"):
        hoja.column_dimensions[col].width = 15
    return hoja


def _tabla_de_meses(hoja_ficha, libro):
    """Rehace la tabla de meses de la Ficha a partir de las hojas que hay."""
    # Se limpia desde FILA_LIMPIEZA y no desde la cabecera actual, porque las
    # primeras versiones ponian la tabla dos filas mas arriba.
    for fila in range(FILA_LIMPIEZA, hoja_ficha.max_row + 2):
        for col in range(1, 7):
            hoja_ficha.cell(row=fila, column=col).value = None

    cabeceras = ["Mes", "Horas", "Días", "A revisar", "Horas extra", "Total a pagar"]
    for i, c in enumerate(cabeceras, start=1):
        celda = hoja_ficha.cell(row=FILA_TABLA_MESES, column=i, value=c)
        celda.font = Font(bold=True)

    meses = sorted((h for h in libro.sheetnames if h != HOJA_FICHA), key=_orden_de_hoja)
    for n, mes in enumerate(meses, start=1):
        fila = FILA_TABLA_MESES + n
        ref = "'{}'".format(mes)
        hoja_ficha.cell(row=fila, column=1, value=mes)
        hoja_ficha.cell(row=fila, column=2, value="={}!$B$3".format(ref)).number_format = "0.00"
        hoja_ficha.cell(row=fila, column=3, value="={}!$B$5".format(ref))
        hoja_ficha.cell(row=fila, column=4, value="={}!$B$4".format(ref))
        hoja_ficha.cell(row=fila, column=5, value="={}!$B$8".format(ref)).number_format = "0.00"
        hoja_ficha.cell(row=fila, column=6, value="={}!$B$6".format(ref)).number_format = EUROS

    if meses:
        fila_total = FILA_TABLA_MESES + len(meses) + 1
        hoja_ficha.cell(row=fila_total, column=1, value="TOTAL").font = Font(bold=True)
        celda = hoja_ficha.cell(
            row=fila_total, column=6,
            value="=SUM(F{}:F{})".format(FILA_TABLA_MESES + 1, FILA_TABLA_MESES + len(meses)))
        celda.font = Font(bold=True)
        celda.number_format = EUROS


def _hoja_mes(libro, periodo, filas):
    """Crea o reemplaza la hoja de un mes. Reemplazar permite rehacer un mes."""
    titulo = nombre_de_mes(periodo)
    if titulo in libro.sheetnames:
        del libro[titulo]
    hoja = libro.create_sheet(titulo)

    completas = [f for f in filas if f["horas"] is not None]
    ultima = PRIMERA_FILA_DATOS + len(filas) - 1

    # El anio no cabe en el nombre de la hoja, asi que va en el titulo de dentro.
    hoja["A1"] = "{} — {} {}".format(filas[0]["empleado"], titulo, periodo[:4])
    hoja["A1"].font = Font(bold=True, size=13)

    hoja["A3"] = "Horas del mes"
    hoja["B3"] = "=ROUND(SUM(D{}:D{}),2)".format(PRIMERA_FILA_DATOS, ultima)
    hoja["B3"].number_format = "0.00"

    hoja["A4"] = "Días a revisar"
    hoja["B4"] = len(filas) - len(completas)

    hoja["A5"] = "Días con horas"
    hoja["B5"] = len(completas)

    sueldo = "{}!$B${}".format(HOJA_FICHA, FILA_SUELDO)
    tipo = "{}!$B${}".format(HOJA_FICHA, FILA_TIPO)
    contrato = "{}!$B${}".format(HOJA_FICHA, FILA_HORAS_CONTRATO)

    hoja["A7"] = "Horas de contrato"
    hoja["B7"] = '=IF({}="{}",{},"")'.format(tipo, PAGO_POR_CONTRATO, contrato)
    hoja["B7"].number_format = "0.00"

    hoja["A8"] = "Horas extra"
    # Puede salir negativo, y se deja verse: significa que esa persona ha hecho
    # menos horas de las que dice su contrato, y es un dato que interesa.
    hoja["B8"] = '=IF({}="{}",ROUND($B$3-{},2),"")'.format(tipo, PAGO_POR_CONTRATO, contrato)
    hoja["B8"].number_format = "0.00"

    hoja["A6"] = "Total a pagar"
    # El sueldo y el tipo de pago se leen de la Ficha, no se repiten en cada mes.
    #   por horas    -> se pagan todas las horas del mes
    #   por contrato -> solo las que exceden las del contrato
    # Las horas extra negativas no restan dinero: se pagan cero y el numero en
    # rojo queda arriba para quien lo tenga que mirar.
    hoja["B6"] = (
        '=IF({sueldo}="","",'
        'IF({tipo}="{contrato_lit}",ROUND(MAX(0,$B$8)*{sueldo},2),'
        'ROUND($B$3*{sueldo},2)))'
    ).format(sueldo=sueldo, tipo=tipo, contrato_lit=PAGO_POR_CONTRATO)
    hoja["B6"].font = Font(bold=True)
    hoja["B6"].number_format = EUROS

    for f in ("A3", "A6", "A8"):
        hoja[f].font = Font(bold=True)

    for i, (cabecera, ancho) in enumerate(CABECERAS, start=1):
        celda = hoja.cell(row=PRIMERA_FILA_DATOS - 1, column=i, value=cabecera)
        celda.font = Font(bold=True)
        hoja.column_dimensions[get_column_letter(i)].width = ancho
    hoja.freeze_panes = "A{}".format(PRIMERA_FILA_DATOS)

    for n, f in enumerate(filas):
        fila = PRIMERA_FILA_DATOS + n
        hoja.cell(row=fila, column=1, value=f["fecha"])
        hoja.cell(row=fila, column=2, value=f["entrada"])
        hoja.cell(row=fila, column=3, value=f["salida"])
        hoja.cell(row=fila, column=4, value=f["horas"]).number_format = "0.00"
        hoja.cell(row=fila, column=5, value=f["aviso"])
        if f["aviso"]:
            for col in range(1, len(CABECERAS) + 1):
                hoja.cell(row=fila, column=col).fill = AMARILLO

    return hoja


def actualizar_libro(contenido, periodo, filas):
    """Devuelve el libro con la hoja del mes anadida o reemplazada.

    contenido es el fichero que ya hay en OneDrive, o None la primera vez.
    """
    if contenido:
        libro = load_workbook(io.BytesIO(contenido))
    else:
        libro = Workbook()
        libro.remove(libro.active)

    primera = filas[0]
    _migrar_nombres_antiguos(libro)
    _hoja_mes(libro, periodo, filas)
    ficha = _ficha(libro, primera["empleado"], primera["workno"], primera["departamento"])

    # Ficha primero y los meses en orden.
    orden = [HOJA_FICHA] + sorted((h for h in libro.sheetnames if h != HOJA_FICHA), key=_orden_de_hoja)
    libro._sheets = [libro[h] for h in orden]

    _tabla_de_meses(ficha, libro)

    buffer = io.BytesIO()
    libro.save(buffer)
    return buffer.getvalue()
